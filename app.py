import os, shutil, json, uuid
from datetime import datetime, timedelta, timezone
from functools import wraps
import logging
from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, Response, session, abort, g, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, extract, case
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from flask_wtf.csrf import CSRFProtect, generate_csrf
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

def _style_xl(ws, headers, col_widths=None):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    header_font = Font(name='Calibri', bold=True, size=11, color='1a1a1a')
    header_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = Border(bottom=Side(style='medium'))
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    else:
        for col_idx, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(len(str(h)) + 3, 12)
    ws.auto_filter.ref = ws.dimensions

logging.basicConfig(
    filename=os.path.join(os.path.dirname(os.path.abspath(__file__)), 'pulse.log'),
    level=logging.WARNING,
    format='%(asctime)s %(levelname)s: %(message)s'
)

app = Flask(__name__)

# ── Secret key: env var > .secret file ──
_secret_key_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.secret')
if os.environ.get('PULSE_SECRET_KEY'):
    app.secret_key = os.environ['PULSE_SECRET_KEY']
elif os.path.exists(_secret_key_file):
    app.secret_key = open(_secret_key_file).read().strip()
else:
    key = os.urandom(24).hex()
    with open(_secret_key_file, 'w') as f:
        f.write(key)
    app.secret_key = key

# ── Session hardening ──
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = False
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)
app.config['SESSION_COOKIE_NAME'] = 'pulse_session'

# ── Database ──
_db_dir = os.environ.get('PULSE_DATA_DIR', os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data'))
os.makedirs(_db_dir, exist_ok=True)
db_path = os.path.join(_db_dir, 'supplypulse.db')
os.makedirs(os.path.dirname(db_path), exist_ok=True)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + db_path
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'connect_args': {
        'timeout': 30,
        'check_same_thread': False,
    },
    'pool_pre_ping': True,
    'pool_recycle': 300,
}

db = SQLAlchemy(app)
csrf = CSRFProtect(app)

# ── Theme image upload ──
THEME_UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'uploads', 'themes')
os.makedirs(THEME_UPLOAD_DIR, exist_ok=True)
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024
ALLOWED_THEME_EXT = {'.jpg', '.jpeg', '.png', '.webp'}

limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=[],
    storage_uri='memory://',
)

# ── Inject CSRF token + theme image into ALL Jinja2 templates ──
@app.context_processor
def inject_globals():
    ctx = dict(csrf_token=generate_csrf(), theme_image=None)
    if 'user_id' in session:
        user = db.session.get(Staff, session['user_id'])
        if user and user.preferences:
            try:
                prefs = json.loads(user.preferences)
                if prefs.get('theme_image'):
                    ctx['theme_image'] = url_for('serve_theme_image', filename=prefs['theme_image'])
            except (json.JSONDecodeError, TypeError):
                pass
    return ctx

# ── Enable SQLite WAL mode on startup ──
with app.app_context():
    from sqlalchemy import text
    try:
        db.session.execute(text('PRAGMA journal_mode=WAL'))
        db.session.execute(text('PRAGMA synchronous=NORMAL'))
        db.session.execute(text('PRAGMA busy_timeout=30000'))
        db.session.execute(text('PRAGMA foreign_keys=ON'))
        db.session.commit()
    except Exception:
        pass


class Store(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), unique=True, nullable=False)
    items = db.relationship('Item', backref='store', lazy=True)


class Unit(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), unique=True, nullable=False)
    wards = db.relationship('Ward', backref='unit', lazy=True)


class Ward(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    unit_id = db.Column(db.Integer, db.ForeignKey('unit.id'), nullable=False)


class Section(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), unique=True, nullable=False)


class Item(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(100), nullable=False)
    name = db.Column(db.String(300), nullable=False)
    store_id = db.Column(db.Integer, db.ForeignKey('store.id'), nullable=False)
    stock_quantity = db.Column(db.Integer, default=0)
    critical_level = db.Column(db.Integer, default=5)


class Week(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    week_number = db.Column(db.Integer, nullable=False, unique=True)
    date_range = db.Column(db.String(200))


class Report(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    item_id = db.Column(db.Integer, db.ForeignKey('item.id'), nullable=False)
    week_id = db.Column(db.Integer, db.ForeignKey('week.id'), nullable=False)
    unit_id = db.Column(db.Integer, db.ForeignKey('unit.id'), nullable=False)
    ward_id = db.Column(db.Integer, db.ForeignKey('ward.id'), nullable=False)
    section_id = db.Column(db.Integer, db.ForeignKey('section.id'), nullable=False)
    status = db.Column(db.String(20), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    item = db.relationship('Item', backref='reports')
    week = db.relationship('Week', backref='reports')
    unit = db.relationship('Unit', backref='reports')
    ward = db.relationship('Ward', backref='reports')
    section = db.relationship('Section', backref='reports')


class Equipment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    biomed = db.Column(db.String(100), unique=True, nullable=False)
    name = db.Column(db.String(300), nullable=False)
    status = db.Column(db.String(30), default='WORKING')
    location = db.Column(db.String(300))
    remarks = db.Column(db.String(500))
    last_service_date = db.Column(db.DateTime)
    next_service_date = db.Column(db.DateTime)
    last_scanned_date = db.Column(db.DateTime)
    last_scanned_by = db.Column(db.String(200))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class EquipmentLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    equipment_id = db.Column(db.Integer, db.ForeignKey('equipment.id'), nullable=False)
    action = db.Column(db.String(100))
    scanned_by = db.Column(db.String(200))
    notes = db.Column(db.String(500))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    equipment = db.relationship('Equipment', backref='logs')


class EquipmentMovement(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    equipment_id = db.Column(db.Integer, db.ForeignKey('equipment.id'), nullable=False)
    checked_out_by = db.Column(db.String(200), nullable=False)
    checked_in_by = db.Column(db.String(200))
    source_location = db.Column(db.String(300))
    destination = db.Column(db.String(300), nullable=False)
    purpose = db.Column(db.String(500))
    checked_out_at = db.Column(db.DateTime, default=datetime.utcnow)
    checked_in_at = db.Column(db.DateTime)
    status = db.Column(db.String(20), default='OUT')

    equipment = db.relationship('Equipment', backref='movements')


class Staff(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.String(20), unique=True, nullable=False)
    full_name = db.Column(db.String(200), nullable=False)
    department = db.Column(db.String(200))
    position = db.Column(db.String(200))
    password_hash = db.Column(db.String(300))
    approved = db.Column(db.Boolean, default=False)
    role = db.Column(db.String(20), default='staff')
    default_unit_id = db.Column(db.Integer, db.ForeignKey('unit.id'), nullable=True)
    default_ward_id = db.Column(db.Integer, db.ForeignKey('ward.id'), nullable=True)
    default_store_id = db.Column(db.Integer, db.ForeignKey('store.id'), nullable=True)
    preferences = db.Column(db.Text, default='{}')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        if not self.password_hash:
            return False
        return check_password_hash(self.password_hash, password)


class ActivityLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user = db.Column(db.String(200))
    action = db.Column(db.String(300))
    details = db.Column(db.String(500))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


def _log_activity(user, action, details=''):
    try:
        ip = request.remote_addr or '0.0.0.0'
        ua = request.user_agent.string if request.user_agent else ''
        full_details = f'{details} | IP: {ip} | UA: {ua[:200]}'
    except RuntimeError:
        full_details = details
    log = ActivityLog(user=user, action=action, details=full_details)
    db.session.add(log)
    db.session.commit()


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        staff = db.session.get(Staff, session['user_id'])
        if not staff or staff.role != 'admin':
            abort(403)
        return f(*args, **kwargs)
    return decorated


with app.app_context():
    db.create_all()
    # migrate: add default_unit_id / default_ward_id to Staff if missing
    from sqlalchemy import inspect as sa_inspect
    try:
        inspector = sa_inspect(db.engine)
        cols = [c['name'] for c in inspector.get_columns('staff')]
        with db.engine.connect() as conn:
            if 'default_unit_id' not in cols:
                conn.execute(text("ALTER TABLE staff ADD COLUMN default_unit_id INTEGER REFERENCES unit(id)"))
            if 'default_ward_id' not in cols:
                conn.execute(text("ALTER TABLE staff ADD COLUMN default_ward_id INTEGER REFERENCES ward(id)"))
            if 'preferences' not in cols:
                conn.execute(text("ALTER TABLE staff ADD COLUMN preferences TEXT DEFAULT '{}'"))
            if 'default_store_id' not in cols:
                conn.execute(text("ALTER TABLE staff ADD COLUMN default_store_id INTEGER REFERENCES store(id)"))
            conn.commit()
    except Exception:
        pass
    admin = Staff.query.filter_by(employee_id='ADMIN').first()
    if admin and not admin.role:
        admin.role = 'admin'
        db.session.commit()


@app.before_request
def log_api_requests():
    if request.path.startswith('/api/') and request.method in ('POST', 'PUT', 'DELETE'):
        app.logger.info(f'{request.method} {request.path} from {request.remote_addr}')

@app.before_request
def require_login():
    allowed = ['login', 'register', 'static']
    if request.endpoint not in allowed and 'user_id' not in session:
        return redirect(url_for('login'))


@app.context_processor
def inject_globals():
    current_user = None
    if 'user_id' in session:
        current_user = db.session.get(Staff, session['user_id'])
    return dict(
        stores=Store.query.order_by(Store.name).all(),
        units=Unit.query.order_by(Unit.name).all(),
        sections=Section.query.order_by(Section.name).all(),
        weeks=Week.query.order_by(Week.week_number).all(),
        wards=Ward.query.order_by(Ward.name).all(),
        current_year=datetime.now().year,
        now=datetime.now,
        search_q=request.args.get('q', ''),
        current_user=current_user
    )


@app.route('/login', methods=['GET', 'POST'])
@limiter.limit('10 per minute')
def login():
    if request.method == 'POST':
        eid = request.form.get('employee_id', '').strip().upper()
        password = request.form.get('password', '')
        staff = Staff.query.filter_by(employee_id=eid).first()
        if staff and staff.check_password(password):
            if not staff.approved:
                flash('Your account is pending admin approval', 'warning')
                return redirect(url_for('login'))
            session.permanent = True
            session['user_id'] = staff.id
            session['employee_id'] = staff.employee_id
            session['full_name'] = staff.full_name
            _log_activity('Login', f'User logged in: {staff.employee_id} — {staff.full_name}')
            flash(f'Welcome back, {staff.full_name}', 'success')
            return redirect(url_for('dashboard'))
        flash('Invalid employee ID or password', 'danger')
        return redirect(url_for('login'))
    return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        eid = request.form.get('employee_id', '').strip().upper()
        name = request.form.get('full_name', '').strip()
        position = request.form.get('position', '').strip()
        dept = request.form.get('department', '').strip()
        password = request.form.get('password', '')
        if not eid or not name or not password:
            flash('Employee ID, Full Name, and Password are required', 'danger')
            return redirect(url_for('register'))
        if len(password) < 8:
            flash('Password must be at least 8 characters', 'danger')
            return redirect(url_for('register'))
        if Staff.query.filter_by(employee_id=eid).first():
            flash(f'Employee ID "{eid}" already exists', 'danger')
            return redirect(url_for('register'))
        s = Staff(employee_id=eid, full_name=name, department=dept, position=position)
        s.set_password(password)
        s.approved = False
        db.session.add(s)
        db.session.commit()
        _log_activity('Registration', f'New registration: {eid} — {name} ({position})')
        flash('Account created! Waiting for admin approval.', 'success')
        return redirect(url_for('login'))
    return render_template('register.html')


@app.route('/logout')
def logout():
    _log_activity('Logout', f'User logged out: {session.get("employee_id", "?")}')
    session.clear()
    flash('Logged out', 'success')
    return redirect(url_for('login'))


def _filter_reports(week_id=None, month=None, year=None, from_date=None, to_date=None):
    q = Report.query
    if week_id:
        q = q.filter(Report.week_id == week_id)
    if year:
        q = q.filter(extract('year', Report.created_at) == year)
    if month:
        q = q.filter(extract('month', Report.created_at) == month)
    if from_date:
        fd = datetime.strptime(from_date, '%Y-%m-%d')
        q = q.filter(Report.created_at >= fd)
    if to_date:
        td = datetime.strptime(to_date, '%Y-%m-%d') + timedelta(days=1)
        q = q.filter(Report.created_at < td)
    return q


def _apply_filters(q, week_id=None, month=None, year=None, from_date=None, to_date=None):
    if week_id: q = q.filter(Report.week_id == week_id)
    if year: q = q.filter(extract('year', Report.created_at) == year)
    if month: q = q.filter(extract('month', Report.created_at) == month)
    if from_date:
        fd = datetime.strptime(from_date, '%Y-%m-%d')
        q = q.filter(Report.created_at >= fd)
    if to_date:
        td = datetime.strptime(to_date, '%Y-%m-%d') + timedelta(days=1)
        q = q.filter(Report.created_at < td)
    return q


@app.route('/')
def dashboard():
    week_id = request.args.get('week_id', type=int)
    month = request.args.get('month', type=int)
    year = request.args.get('year', type=int)
    from_date = request.args.get('from')
    to_date = request.args.get('to')

    has_filter = week_id or month or year or from_date or to_date
    fkw = dict(week_id=week_id, month=month, year=year, from_date=from_date, to_date=to_date)
    total_reports = _filter_reports(**fkw).count()
    shortage_count = _filter_reports(**fkw).filter_by(status='Shortage').count()
    una_count = _filter_reports(**fkw).filter_by(status='Not available').count()

    if has_filter:
        item_ids_q = _filter_reports(**fkw).with_entities(Report.item_id).distinct()
        base_q = Item.query.filter(Item.id.in_(item_ids_q))
    else:
        base_q = Item.query

    total_items = base_q.count()

    pending_approvals = Staff.query.filter_by(approved=False).count()
    current_user = db.session.get(Staff, session.get('user_id')) if session.get('user_id') else None

    recent_reports = _filter_reports(**fkw).order_by(
        Report.created_at.desc()).limit(10).all()

    shortage_by_store = _apply_filters(db.session.query(
        Store.name, func.count(Report.id)
    ).select_from(Report).join(Item).join(Store).filter(
        Report.status == 'Shortage'
    ), **fkw).group_by(Store.name).all()

    una_by_store = _apply_filters(db.session.query(
        Store.name, func.count(Report.id)
    ).select_from(Report).join(Item).join(Store).filter(
        Report.status == 'Not available'
    ), **fkw).group_by(Store.name).all()

    shortage_by_week = _apply_filters(db.session.query(
        Week.week_number, Week.date_range, func.count(Report.id)
    ).select_from(Report).join(Week).filter(
        Report.status == 'Shortage'
    ), **fkw).group_by(Week.week_number, Week.date_range).order_by(Week.week_number).all()

    una_by_week = _apply_filters(db.session.query(
        Week.week_number, Week.date_range, func.count(Report.id)
    ).select_from(Report).join(Week).filter(
        Report.status == 'Not available'
    ), **fkw).group_by(Week.week_number, Week.date_range).order_by(Week.week_number).all()

    critical_items = base_q.filter(
        Item.stock_quantity > 0,
        Item.stock_quantity <= Item.critical_level
    ).order_by(Item.stock_quantity).all()

    out_of_stock = base_q.filter(
        Item.stock_quantity == 0
    ).order_by(Item.name).all()

    ok_count = base_q.filter(
        Item.stock_quantity > Item.critical_level
    ).count()
    low_count = len(critical_items)
    critical_count = len(out_of_stock)

    crit_total = low_count + critical_count
    noncrit_total = ok_count

    return render_template('index.html',
                           total_items=total_items,
                           total_reports=total_reports,
                           shortage_count=shortage_count,
                           una_count=una_count,
                           recent_reports=recent_reports,
                           shortage_by_store=shortage_by_store,
                           una_by_store=una_by_store,
                           shortage_by_week=shortage_by_week,
                           una_by_week=una_by_week,
                           critical_items=critical_items,
                           out_of_stock=out_of_stock,
                           ok_count=ok_count,
                           low_count=low_count,
                           critical_count=critical_count,
                           crit_total=crit_total,
                           noncrit_total=noncrit_total,
                           from_date=from_date,
                           to_date=to_date,
                           weeks=Week.query.order_by(Week.week_number).all(),
                           sel_week=week_id,
                           sel_month=month,
                           sel_year=year,
                            now=datetime.now(timezone.utc),
                            pending_approvals=pending_approvals,
                            current_user=current_user)


@app.route('/api/dashboard')
def api_dashboard():
    week_id = request.args.get('week_id', type=int)
    month = request.args.get('month', type=int)
    year = request.args.get('year', type=int)
    from_date = request.args.get('from')
    to_date = request.args.get('to')

    has_filter = week_id or month or year or from_date or to_date
    fkw = dict(week_id=week_id, month=month, year=year, from_date=from_date, to_date=to_date)

    total_reports = _filter_reports(**fkw).count()
    shortage_count = _filter_reports(**fkw).filter_by(status='Shortage').count()
    una_count = _filter_reports(**fkw).filter_by(status='Not available').count()

    if has_filter:
        item_ids_q = _filter_reports(**fkw).with_entities(Report.item_id).distinct()
        base_q = Item.query.filter(Item.id.in_(item_ids_q))
    else:
        base_q = Item.query

    total_items = base_q.count()
    out_of_stock = base_q.filter(Item.stock_quantity == 0).order_by(Item.name).all()
    critical_items = base_q.filter(
        Item.stock_quantity > 0,
        Item.stock_quantity <= Item.critical_level
    ).order_by(Item.stock_quantity).all()
    ok_count = base_q.filter(Item.stock_quantity > Item.critical_level).count()
    low_count = len(critical_items)
    critical_count = len(out_of_stock)

    shortage_by_store = _apply_filters(db.session.query(
        Store.name, func.count(Report.id)
    ).select_from(Report).join(Item).join(Store).filter(
        Report.status == 'Shortage'
    ), **fkw).group_by(Store.name).all()

    una_by_store = _apply_filters(db.session.query(
        Store.name, func.count(Report.id)
    ).select_from(Report).join(Item).join(Store).filter(
        Report.status == 'Not available'
    ), **fkw).group_by(Store.name).all()

    shortage_by_week = _apply_filters(db.session.query(
        Week.week_number, Week.date_range, func.count(Report.id)
    ).select_from(Report).join(Week).filter(
        Report.status == 'Shortage'
    ), **fkw).group_by(Week.week_number, Week.date_range).order_by(Week.week_number).all()

    una_by_week = _apply_filters(db.session.query(
        Week.week_number, Week.date_range, func.count(Report.id)
    ).select_from(Report).join(Week).filter(
        Report.status == 'Not available'
    ), **fkw).group_by(Week.week_number, Week.date_range).order_by(Week.week_number).all()

    recent = _filter_reports(**fkw).order_by(Report.created_at.desc()).limit(10).all()
    recent_data = []
    for r in recent:
        recent_data.append({
            'item': f'{r.item.code} - {r.item.name}',
            'store': r.item.store.name,
            'unit': r.unit.name,
            'ward': r.ward.name,
            'status': r.status,
            'week': f'Week {r.week.week_number}',
            'date': r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else '',
        })

    crit_total = low_count + critical_count
    noncrit_total = ok_count

    return jsonify({
        'total_items': total_items,
        'total_reports': total_reports,
        'shortage_count': shortage_count,
        'una_count': una_count,
        'crit_total': crit_total,
        'noncrit_total': noncrit_total,
        'ok_count': ok_count,
        'low_count': low_count,
        'critical_count': critical_count,
        'critical_items': [{'id': i.id, 'code': i.code, 'name': i.name, 'stock': i.stock_quantity, 'critical': i.critical_level, 'store': i.store.name} for i in critical_items],
        'out_of_stock': [{'id': i.id, 'code': i.code, 'name': i.name, 'stock': i.stock_quantity, 'critical': i.critical_level, 'store': i.store.name} for i in out_of_stock],
        'shortage_by_store': [{'store': s[0], 'count': s[1]} for s in shortage_by_store],
        'una_by_store': [{'store': s[0], 'count': s[1]} for s in una_by_store],
        'shortage_by_week': [{'week': f'Week {w[0]}', 'count': w[2]} for w in shortage_by_week],
        'una_by_week': [{'week': f'Week {w[0]}', 'count': w[2]} for w in una_by_week],
        'recent': recent_data,
    })


@app.route('/api/alerts')
def api_alerts():
    critical = Item.query.filter(
        Item.stock_quantity > 0,
        Item.stock_quantity <= Item.critical_level
    ).order_by(Item.stock_quantity).all()
    out_of_stock = Item.query.filter(
        Item.stock_quantity == 0
    ).order_by(Item.name).all()
    return jsonify({
        'total': len(critical) + len(out_of_stock),
        'critical_count': len(critical),
        'out_count': len(out_of_stock),
        'items': (
            [{'id': i.id, 'code': i.code, 'name': i.name, 'stock': i.stock_quantity, 'critical': i.critical_level, 'store': i.store.name, 'type': 'low'} for i in critical] +
            [{'id': i.id, 'code': i.code, 'name': i.name, 'stock': 0, 'critical': i.critical_level, 'store': i.store.name, 'type': 'out'} for i in out_of_stock]
        )
    })


@app.route('/api/trends')
def api_trends():
    month = request.args.get('month', type=int)
    year = request.args.get('year', type=int)
    from_date = request.args.get('from')
    to_date = request.args.get('to')

    q = db.session.query(
        extract('year', Report.created_at).label('year'),
        extract('month', Report.created_at).label('month'),
        func.count(Report.id).label('total'),
        func.sum(case((Report.status == 'Shortage', 1), else_=0)).label('shortage'),
        func.sum(case((Report.status == 'Not available', 1), else_=0)).label('una'),
    )

    has_filter = month or year or from_date or to_date
    if has_filter:
        if year:
            q = q.filter(extract('year', Report.created_at) == year)
        if month:
            q = q.filter(extract('month', Report.created_at) == month)
        if from_date:
            fd = datetime.strptime(from_date, '%Y-%m-%d')
            q = q.filter(Report.created_at >= fd)
        if to_date:
            td = datetime.strptime(to_date, '%Y-%m-%d') + timedelta(days=1)
            q = q.filter(Report.created_at < td)
    else:
        q = q.filter(Report.created_at >= datetime.now() - timedelta(days=365))

    rows = q.group_by('year', 'month').order_by('year', 'month').all()
    return jsonify([{
        'label': f'{int(r.year)}-{int(r.month):02d}',
        'total': r.total,
        'shortage': int(r.shortage),
        'una': int(r.una),
    } for r in rows])


@app.route('/items/<int:id>/history')
@login_required
def item_history(id):
    item = db.session.get(Item, id)
    if not item:
        flash('Item not found', 'danger')
        return redirect(url_for('items'))
    reports = Report.query.filter_by(item_id=id).order_by(Report.created_at.desc()).all()
    chart_data = db.session.query(
        func.strftime('%Y-%m', Report.created_at).label('month'),
        func.count(Report.id).label('total'),
        func.sum(case((Report.status == 'Shortage', 1), else_=0)).label('shortage'),
        func.sum(case((Report.status == 'Not available', 1), else_=0)).label('una'),
    ).filter(Report.item_id == id).group_by('month').order_by('month').all()
    return render_template('item_history.html', item=item, reports=reports, chart_data=chart_data)


@app.route('/reorder')
@login_required
def reorder():
    ninety_days_ago = datetime.now() - timedelta(days=90)
    suggestions = db.session.query(
        Item.id, Item.code, Item.name, Item.stock_quantity, Item.critical_level,
        Store.name.label('store'),
        func.count(Report.id).label('incidents'),
        func.sum(case((Report.status == 'Shortage', 1), else_=0)).label('shortages'),
        func.sum(case((Report.status == 'Not available', 1), else_=0)).label('unas'),
    ).select_from(Report).join(Item).join(Store).filter(
        Report.created_at >= ninety_days_ago
    ).group_by(Item.id, Item.code, Item.name, Item.stock_quantity, Item.critical_level, Store.name).order_by(
        func.count(Report.id).desc()
    ).all()
    reorder_list = []
    for s in suggestions:
        incidents = s.incidents or 0
        if incidents >= 2:
            reorder_amt = max(s.critical_level or 10, incidents * 2)
            reorder_list.append({
                'id': s.id, 'code': s.code, 'name': s.name,
                'stock': s.stock_quantity or 0, 'critical': s.critical_level or 0,
                'store': s.store, 'incidents': s.incidents or 0,
                'shortages': s.shortages or 0, 'unas': s.unas or 0,
                'suggested': reorder_amt,
            })
    return render_template('reorder.html', suggestions=reorder_list)


@app.route('/api/reorder/restock/<int:id>', methods=['POST'])
@login_required
def reorder_restock(id):
    item = db.session.get(Item, id)
    if not item:
        return jsonify({'error': 'Item not found'}), 404
    data = request.get_json()
    if not data or 'quantity' not in data:
        return jsonify({'error': 'Missing quantity'}), 400
    qty = int(data['quantity'])
    if qty < 1:
        return jsonify({'error': 'Quantity must be at least 1'}), 400
    old_qty = item.stock_quantity or 0
    item.stock_quantity = old_qty + qty
    db.session.commit()
    user = session.get('employee_id', '?')
    _log_activity(user, 'Restock', f'Item #{item.id} {item.code} — {item.name}: {old_qty} → {item.stock_quantity} (+{qty})')
    return jsonify({
        'id': item.id,
        'stock': item.stock_quantity,
        'critical': item.critical_level,
        'ok': item.stock_quantity > item.critical_level,
    })


@app.route('/preferences', methods=['GET', 'POST'])
@login_required
def preferences():
    user = db.session.get(Staff, session['user_id'])
    units = Unit.query.order_by(Unit.name).all()
    wards = Ward.query.order_by(Ward.name).all()

    if request.method == 'POST':
        user.default_unit_id = request.form.get('default_unit_id', type=int) or None
        user.default_ward_id = request.form.get('default_ward_id', type=int) or None
        user.default_store_id = request.form.get('default_store_id', type=int) or None
        pref_theme = request.form.get('pref_theme', 'auto')
        pref_poll = request.form.get('pref_poll', type=int) or 10
        try:
            prefs = json.loads(user.preferences) if user.preferences else {}
        except (json.JSONDecodeError, TypeError):
            prefs = {}
        prefs['theme'] = pref_theme
        prefs['poll_interval'] = pref_poll
        prefs['auto_redirect'] = 'auto_redirect' in request.form
        user.preferences = json.dumps(prefs)
        db.session.commit()
        if pref_theme != 'auto':
            resp = redirect(url_for('preferences'))
            resp.set_cookie('theme', pref_theme, max_age=31536000, path='/')
        else:
            resp = redirect(url_for('preferences'))
            resp.delete_cookie('theme', path='/')
        flash('Preferences saved', 'success')
        return resp

    try:
        current_prefs = json.loads(user.preferences) if user.preferences else {}
    except (json.JSONDecodeError, TypeError):
        current_prefs = {}
    stores = Store.query.order_by(Store.name).all()
    theme_img = None
    try:
        prefs = json.loads(user.preferences) if user.preferences else {}
    except (json.JSONDecodeError, TypeError):
        prefs = {}
    if prefs.get('theme_image'):
        theme_img = prefs['theme_image']
    return render_template('preferences.html', user=user, units=units, wards=wards, stores=stores, prefs=current_prefs, theme_img=theme_img)


@app.route('/preferences/upload-theme', methods=['POST'])
@login_required
def upload_theme():
    user = db.session.get(Staff, session['user_id'])
    file = request.files.get('theme_image')
    if not file or not file.filename:
        flash('No file selected', 'danger')
        return redirect(url_for('preferences'))
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALLOWED_THEME_EXT:
        flash('Allowed formats: JPG, PNG, WebP', 'danger')
        return redirect(url_for('preferences'))
    filename = f'user_{user.id}_{uuid.uuid4().hex[:12]}{ext}'
    file.save(os.path.join(THEME_UPLOAD_DIR, filename))
    try:
        prefs = json.loads(user.preferences) if user.preferences else {}
    except (json.JSONDecodeError, TypeError):
        prefs = {}
    old = prefs.get('theme_image')
    if old:
        old_path = os.path.join(THEME_UPLOAD_DIR, os.path.basename(old))
        if os.path.exists(old_path):
            os.remove(old_path)
    prefs['theme_image'] = filename
    user.preferences = json.dumps(prefs)
    db.session.commit()
    flash('Theme image uploaded', 'success')
    return redirect(url_for('preferences'))


@app.route('/preferences/remove-theme', methods=['POST'])
@login_required
def remove_theme():
    user = db.session.get(Staff, session['user_id'])
    try:
        prefs = json.loads(user.preferences) if user.preferences else {}
    except (json.JSONDecodeError, TypeError):
        prefs = {}
    old = prefs.pop('theme_image', None)
    if old:
        old_path = os.path.join(THEME_UPLOAD_DIR, os.path.basename(old))
        if os.path.exists(old_path):
            os.remove(old_path)
    user.preferences = json.dumps(prefs)
    db.session.commit()
    flash('Theme image removed', 'success')
    return redirect(url_for('preferences'))


@app.route('/items')
def items():
    all_items = Item.query.order_by(Item.code).all()
    stores = Store.query.order_by(Store.name).all()
    return render_template('items.html', items=all_items, stores=stores)


@app.route('/items/add', methods=['POST'])
def add_item():
    code = request.form.get('code', '').strip()
    name = request.form.get('name', '').strip()
    store_id = request.form.get('store_id')
    if code and name and store_id:
        item = Item(code=code, name=name, store_id=store_id)
        sq = request.form.get('stock_quantity', '')
        cl = request.form.get('critical_level', '')
        if sq:
            item.stock_quantity = int(sq)
        if cl:
            item.critical_level = int(cl)
        db.session.add(item)
        db.session.commit()
        flash('Item added successfully', 'success')
    return redirect(url_for('items'))


@app.route('/items/<int:id>/edit', methods=['POST'])
def edit_item(id):
    item = db.session.get(Item, id)
    if item:
        item.code = request.form.get('code', item.code).strip()
        item.name = request.form.get('name', item.name).strip()
        item.store_id = int(request.form.get('store_id', item.store_id))
        sq = request.form.get('stock_quantity', '')
        cl = request.form.get('critical_level', '')
        if sq:
            item.stock_quantity = int(sq)
        if cl:
            item.critical_level = int(cl)
        db.session.commit()
        flash('Item updated', 'success')
    return redirect(url_for('items'))


@app.route('/items/<int:id>/delete', methods=['POST'])
def delete_item(id):
    item = db.session.get(Item, id)
    if item:
        Report.query.filter_by(item_id=id).delete()
        db.session.delete(item)
        db.session.commit()
        flash('Item deleted', 'success')
    return redirect(url_for('items'))


@app.route('/items/import', methods=['POST'])
def import_items():
    import openpyxl
    file = request.files.get('file')
    if not file:
        flash('No file selected', 'danger')
        return redirect(url_for('items'))
    wb = openpyxl.load_workbook(file)
    ws = wb['Master List2'] if 'Master List2' in wb.sheetnames else wb.active

    code_idx = None
    name_idx = None
    store_idx = None
    stock_idx = None
    cl_idx = None
    data_start = 1
    for row in ws.iter_rows(min_row=1, max_row=5, max_col=ws.min_column + 20, values_only=True):
        vals = [str(v).strip().lower() if v else '' for v in row]
        if 'code' in vals:
            code_idx = vals.index('code')
            name_idx = vals.index('item') if 'item' in vals else None
            store_idx = vals.index('store') if 'store' in vals else None
            stock_idx = vals.index('stock quantity') if 'stock quantity' in vals else None
            cl_idx = vals.index('critical level') if 'critical level' in vals else None
            data_start = data_start + 1
            break
        data_start += 1
    if code_idx is None:
        try:
            first = next(ws.iter_rows(min_row=data_start, max_row=data_start, max_col=data_start + 10, values_only=True))
            idx0 = str(first[0]).strip() if first[0] is not None else ''
        except:
            idx0 = ''
        if idx0.isdigit() and data_start > 1:
            code_idx, name_idx, store_idx = 1, 2, 3
        else:
            code_idx, name_idx, store_idx = 0, 1, 2

    need = [i for i in (code_idx, name_idx, store_idx) if i is not None]
    max_data_col = max([i for i in (code_idx, name_idx, store_idx, stock_idx, cl_idx) if i is not None]) + 1 + ws.min_column

    count = 0
    for row in ws.iter_rows(min_row=data_start, min_col=ws.min_column, max_col=max_data_col, values_only=True):
        vals = [str(v).strip() if v is not None else '' for v in row]
        if len(vals) <= max(need):
            continue
        code = vals[code_idx]
        name = vals[name_idx] if name_idx is not None and len(vals) > name_idx else ''
        store_name = vals[store_idx] if store_idx is not None and len(vals) > store_idx else ''
        if not code or not name or not store_name:
            continue
        store = Store.query.filter_by(name=store_name).first()
        if not store:
            store = Store(name=store_name)
            db.session.add(store)
            db.session.commit()
        if not Item.query.filter_by(code=code).first():
            item = Item(code=code, name=name, store_id=store.id)
            if stock_idx is not None and len(vals) > stock_idx and vals[stock_idx]:
                try: item.stock_quantity = int(float(vals[stock_idx]))
                except: pass
            if cl_idx is not None and len(vals) > cl_idx and vals[cl_idx]:
                try: item.critical_level = int(float(vals[cl_idx]))
                except: pass
            db.session.add(item)
            count += 1
    db.session.commit()
    flash(f'{count} items imported', 'success')
    return redirect(url_for('items'))


@app.route('/report/new')
def report_new():
    items_q = Item.query.order_by(Item.code).all()
    items = [{'id': i.id, 'code': i.code, 'name': i.name,
              'store_name': i.store.name if i.store else ''} for i in items_q]
    weeks = Week.query.order_by(Week.week_number).all()
    units = Unit.query.order_by(Unit.name).all()
    sections = Section.query.order_by(Section.name).all()
    today = datetime.now(timezone.utc).isocalendar()
    current_week_number = today[1]
    current_week = Week.query.filter_by(week_number=current_week_number).first()
    current_user = db.session.get(Staff, session.get('user_id')) if session.get('user_id') else None
    default_unit_id = current_user.default_unit_id if current_user else None
    default_ward_id = current_user.default_ward_id if current_user else None
    wards = []
    if default_unit_id:
        wards = Ward.query.filter_by(unit_id=default_unit_id).order_by(Ward.name).all()
    return render_template('report_new.html', items=items, weeks=weeks,
                           units=units, sections=sections,
                           current_week_id=current_week.id if current_week else None,
                           default_unit_id=default_unit_id,
                           default_ward_id=default_ward_id,
                           default_wards=wards)


@app.route('/api/report/quick', methods=['POST'])
def api_report_quick():
    item_id = request.form.get('item_id')
    if not item_id:
        return jsonify({'error': 'item_id required'}), 400
    current_user = db.session.get(Staff, session.get('user_id')) if session.get('user_id') else None
    unit_id = (current_user.default_unit_id if current_user and current_user.default_unit_id
               else request.form.get('unit_id'))
    ward_id = (current_user.default_ward_id if current_user and current_user.default_ward_id
               else request.form.get('ward_id'))
    if not unit_id or not ward_id:
        ward = Ward.query.first()
        unit_id = ward.unit_id if ward else 1
        ward_id = ward.id if ward else 1
    today = datetime.now(timezone.utc).isocalendar()
    week = Week.query.filter_by(week_number=today[1]).first()
    if not week:
        flash('No active week found', 'danger')
        return redirect(url_for('dashboard'))
    section = Section.query.first()
    report = Report(
        item_id=int(item_id), week_id=week.id,
        unit_id=int(unit_id), ward_id=int(ward_id),
        section_id=section.id if section else 1,
        status='Shortage'
    )
    db.session.add(report)
    db.session.commit()
    flash('Shortage reported!', 'success')
    return redirect(url_for('dashboard'))


@app.route('/api/wards/<int:unit_id>')
def api_wards(unit_id):
    wards = Ward.query.filter_by(unit_id=unit_id).order_by(Ward.name).all()
    return jsonify([{'id': w.id, 'name': w.name} for w in wards])


@app.route('/report/submit', methods=['POST'])
def report_submit():
    item_id = request.form.get('item_id')
    week_id = request.form.get('week_id')
    unit_id = request.form.get('unit_id')
    ward_id = request.form.get('ward_id')
    section_id = request.form.get('section_id')
    status = request.form.get('status')
    if all([item_id, week_id, unit_id, ward_id, section_id, status]):
        report = Report(
            item_id=item_id, week_id=week_id, unit_id=unit_id,
            ward_id=ward_id, section_id=section_id, status=status
        )
        db.session.add(report)
        db.session.commit()
        flash('Report submitted', 'success')
    else:
        flash('All fields required', 'danger')
    return redirect(url_for('reports'))


@app.route('/report/import', methods=['POST'])
def import_reports():
    import openpyxl
    file = request.files.get('file')
    if not file:
        flash('No file selected', 'danger')
        return redirect(url_for('reports'))
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb['Report'] if 'Report' in wb.sheetnames else wb.active

        header = None
        header_row = 0
        for row in ws.iter_rows(min_row=1, max_row=5, max_col=ws.min_column + 20, values_only=True):
            header_row += 1
            vals = [str(v).strip().lower() if v else '' for v in row]
            if 'code' in vals:
                header = vals
                break
        if not header:
            flash('Could not find header row with "Code" column in Report sheet', 'danger')
            return redirect(url_for('reports'))
        try:
            code_idx = header.index('code')
            status_idx = header.index('status')
        except ValueError as e:
            flash(f'Missing required column in Report sheet: {e}', 'danger')
            return redirect(url_for('reports'))
        name_idx = header.index('item') if 'item' in header else None
        store_idx = header.index('store') if 'store' in header else None
        week_idx = header.index('week') if 'week' in header else None
        unit_idx = header.index('unit') if 'unit' in header else None
        ward_idx = header.index('ward') if 'ward' in header else None
        section_idx = header.index('section') if 'section' in header else None

        need = [i for i in (code_idx, status_idx, name_idx, store_idx, week_idx, unit_idx, ward_idx, section_idx) if i is not None]
        max_data_col = max(need) + 1 + ws.min_column

        count = 0
        errors = []
        for i, row in enumerate(ws.iter_rows(min_row=header_row + 1, min_col=ws.min_column, max_col=max_data_col, values_only=True), start=header_row + 1):
            vals = [str(v).strip() if v is not None else '' for v in row]
            if len(vals) <= max(code_idx, status_idx):
                continue
            code = vals[code_idx]
            status_raw = vals[status_idx]
            if not code or not status_raw:
                continue
            section_name = vals[section_idx] if section_idx is not None and len(vals) > section_idx else ''
            unit_name = vals[unit_idx] if unit_idx is not None and len(vals) > unit_idx else ''
            ward_name = vals[ward_idx] if ward_idx is not None and len(vals) > ward_idx else ''
            item_name = vals[name_idx] if name_idx is not None and len(vals) > name_idx else ''
            store_name = vals[store_idx] if store_idx is not None and len(vals) > store_idx else ''
            week_raw = vals[week_idx] if week_idx is not None and len(vals) > week_idx else ''
            item = Item.query.filter_by(code=code).first()
            if not item:
                errors.append(f'Row {i}: Item code "{code}" not found')
                continue
            import re
            nums = re.findall(r'\d+', str(week_raw))
            week_num = int(nums[0]) if nums else None
            week = Week.query.filter_by(week_number=week_num).first() if week_num else None
            if not week and week_num:
                week = Week(week_number=week_num, date_range=str(week_raw))
                db.session.add(week)
                db.session.commit()
            if not week:
                errors.append(f'Row {i}: Week not found (value: {week_raw})')
                continue
            unit = Unit.query.filter_by(name=unit_name).first() if unit_name else None
            ward = Ward.query.filter_by(name=ward_name).first() if ward_name else None
            section = Section.query.filter_by(name=section_name).first() if section_name else None
            if not unit:
                unit = Unit(name=unit_name) if unit_name else None
                if unit:
                    db.session.add(unit)
                    db.session.commit()
            if not ward and unit and ward_name:
                ward = Ward(name=ward_name, unit_id=unit.id)
                db.session.add(ward)
                db.session.commit()
            if not section and section_name:
                section = Section(name=section_name)
                db.session.add(section)
                db.session.commit()
            s = 'Shortage' if status_raw.lower() == 'shortage' else 'Not available'
            report = Report(
                item_id=item.id, week_id=week.id,
                unit_id=unit.id if unit else 1,
                ward_id=ward.id if ward else 1,
                section_id=section.id if section else 1,
                status=s
            )
            db.session.add(report)
            count += 1
        db.session.commit()
        msg = f'{count} reports imported'
        if errors:
            msg += f'. Errors: {"; ".join(errors[:5])}'
        flash(msg, 'success' if count else 'danger')
    except Exception as e:
        flash(f'Import failed: {e}', 'danger')
    return redirect(url_for('reports'))


@app.route('/reports')
def reports():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    if per_page not in (25, 50, 100, 200):
        per_page = 50
    store_id = request.args.get('store_id', type=int)
    status_filter = request.args.get('status')
    week_id = request.args.get('week_id', type=int)
    from_date = request.args.get('from')
    to_date = request.args.get('to')

    query = Report.query
    if store_id:
        query = query.join(Item).filter(Item.store_id == store_id)
    if status_filter:
        query = query.filter(Report.status == status_filter)
    if week_id:
        query = query.filter(Report.week_id == week_id)
    if from_date:
        fd = datetime.strptime(from_date, '%Y-%m-%d')
        query = query.filter(Report.created_at >= fd)
    if to_date:
        td = datetime.strptime(to_date, '%Y-%m-%d') + timedelta(days=1)
        query = query.filter(Report.created_at < td)

    query = query.order_by(Report.created_at.desc())
    pagination = query.paginate(page=page, per_page=per_page)
    return render_template('reports.html', pagination=pagination,
                           per_page=per_page,
                           stores=Store.query.order_by(Store.name).all(),
                           from_date=from_date,
                           to_date=to_date)


@app.route('/shortage')
def shortage():
    store_id = request.args.get('store_id', type=int)
    from_date = request.args.get('from')
    to_date = request.args.get('to')

    query = db.session.query(
        Item.code, Item.name, Store.name.label('store'),
        func.count(Report.id).label('total')
    ).select_from(Report).join(Item).join(Store).filter(
        Report.status == 'Shortage'
    )
    if store_id:
        query = query.filter(Item.store_id == store_id)
    if from_date:
        fd = datetime.strptime(from_date, '%Y-%m-%d')
        query = query.filter(Report.created_at >= fd)
    if to_date:
        td = datetime.strptime(to_date, '%Y-%m-%d') + timedelta(days=1)
        query = query.filter(Report.created_at < td)
    data = query.group_by(Item.id, Item.code, Item.name, Store.name)\
                .order_by(func.count(Report.id).desc()).all()

    return render_template('shortage.html', data=data,
                           from_date=from_date, to_date=to_date,
                           stores=Store.query.order_by(Store.name).all())


@app.route('/una')
def una():
    store_id = request.args.get('store_id', type=int)
    from_date = request.args.get('from')
    to_date = request.args.get('to')

    query = db.session.query(
        Item.code, Item.name, Store.name.label('store'),
        func.count(Report.id).label('total')
    ).select_from(Report).join(Item).join(Store).filter(
        Report.status == 'Not available'
    )
    if store_id:
        query = query.filter(Item.store_id == store_id)
    if from_date:
        fd = datetime.strptime(from_date, '%Y-%m-%d')
        query = query.filter(Report.created_at >= fd)
    if to_date:
        td = datetime.strptime(to_date, '%Y-%m-%d') + timedelta(days=1)
        query = query.filter(Report.created_at < td)
    data = query.group_by(Item.id, Item.code, Item.name, Store.name)\
                .order_by(func.count(Report.id).desc()).all()

    return render_template('una.html', data=data,
                           from_date=from_date, to_date=to_date,
                           stores=Store.query.order_by(Store.name).all())


@app.route('/settings')
@admin_required
def settings():
    stores = Store.query.order_by(Store.name).all()
    units = Unit.query.order_by(Unit.name).all()
    wards = Ward.query.order_by(Ward.name).all()
    sections = Section.query.order_by(Section.name).all()
    weeks = Week.query.order_by(Week.week_number).all()
    return render_template('settings.html', stores=stores, units=units,
                           wards=wards, sections=sections, weeks=weeks)


@app.route('/settings/store/add', methods=['POST'])
def add_store():
    name = request.form.get('name', '').strip()
    if name and not Store.query.filter_by(name=name).first():
        db.session.add(Store(name=name))
        db.session.commit()
        _log_activity('Settings', f'Added store: {name}')
        flash('Store added', 'success')
    return redirect(url_for('settings'))


@app.route('/settings/store/<int:id>/delete', methods=['POST'])
def delete_store(id):
    store = db.session.get(Store, id)
    if store:
        Item.query.filter_by(store_id=id).delete()
        db.session.delete(store)
        db.session.commit()
        flash('Store deleted', 'success')
    return redirect(url_for('settings'))


@app.route('/settings/unit/add', methods=['POST'])
def add_unit():
    name = request.form.get('name', '').strip()
    if name and not Unit.query.filter_by(name=name).first():
        db.session.add(Unit(name=name))
        db.session.commit()
        _log_activity('Settings', f'Added unit: {name}')
        flash('Unit added', 'success')
    return redirect(url_for('settings'))


@app.route('/settings/ward/add', methods=['POST'])
def add_ward():
    name = request.form.get('name', '').strip()
    unit_id = request.form.get('unit_id')
    if name and unit_id:
        db.session.add(Ward(name=name, unit_id=unit_id))
        db.session.commit()
        _log_activity('Settings', f'Added ward: {name}')
        flash('Ward added', 'success')
    return redirect(url_for('settings'))


@app.route('/settings/section/add', methods=['POST'])
def add_section():
    name = request.form.get('name', '').strip()
    if name and not Section.query.filter_by(name=name).first():
        db.session.add(Section(name=name))
        db.session.commit()
        flash('Section added', 'success')
    return redirect(url_for('settings'))


@app.route('/settings/week/add', methods=['POST'])
def add_week():
    week_number = request.form.get('week_number', type=int)
    date_range = request.form.get('date_range', '').strip()
    if week_number and not Week.query.filter_by(week_number=week_number).first():
        db.session.add(Week(week_number=week_number, date_range=date_range))
        db.session.commit()
        _log_activity('Settings', f'Added week: {week_number}')
        flash('Week added', 'success')
    return redirect(url_for('settings'))


@app.route('/settings/seed', methods=['POST'])
def seed_from_excel():
    import openpyxl
    file = request.files.get('file')
    if not file:
        flash('No file selected', 'danger')
        return redirect(url_for('settings'))

    wb = openpyxl.load_workbook(file)

    if 'Droplist' in wb.sheetnames:
        ws = wb['Droplist']
        mc = ws.min_column
        if mc < 1:
            mc = 1

        for row in ws.iter_rows(min_row=2, min_col=mc, max_col=mc + 15, values_only=True):
            vals = list(row)
            date_raw = str(vals[0]).strip() if len(vals) > 0 and vals[0] else ''
            week_raw = vals[1] if len(vals) > 1 else None
            if isinstance(week_raw, (int, float)) and week_raw > 0:
                if not Week.query.filter_by(week_number=int(week_raw)).first():
                    db.session.add(Week(week_number=int(week_raw), date_range=date_raw))

        store_names = set()
        for row in ws.iter_rows(min_row=2, min_col=mc, max_col=mc + 15, values_only=True):
            vals = list(row)
            store_val = str(vals[3]).strip() if len(vals) > 3 and vals[3] else ''
            if store_val:
                store_names.add(store_val)
        for name in store_names:
            if not Store.query.filter_by(name=name).first():
                db.session.add(Store(name=name))

        rows_data = []
        for row in ws.iter_rows(min_row=5, min_col=mc, max_col=mc + 15, values_only=True):
            vals = list(row)
            u = str(vals[7]).strip() if len(vals) > 7 and vals[7] else ''
            w = str(vals[8]).strip() if len(vals) > 8 and vals[8] else ''
            s = str(vals[9]).strip() if len(vals) > 9 and vals[9] else ''
            if u or s:
                rows_data.append((u, w, s))

        seen_units = set()
        for u, w, s in rows_data:
            if u and u not in seen_units:
                seen_units.add(u)
                if not Unit.query.filter_by(name=u).first():
                    db.session.add(Unit(name=u))
            if s and not Section.query.filter_by(name=s).first():
                db.session.add(Section(name=s))

        db.session.commit()

        for u, w, s in rows_data:
            if u and w:
                unit = Unit.query.filter_by(name=u).first()
                if unit and not Ward.query.filter_by(name=w, unit_id=unit.id).first():
                    db.session.add(Ward(name=w, unit_id=unit.id))

        db.session.commit()

    flash('Reference data seeded from Excel', 'success')
    return redirect(url_for('settings'))


@app.route('/import-all', methods=['POST'])
def import_all():
    import openpyxl
    file = request.files.get('file')
    if not file:
        flash('No file selected', 'danger')
        return redirect(request.referrer or url_for('dashboard'))

    wb = openpyxl.load_workbook(file)
    results = []

    def _find_header(ws, keywords):
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, max_col=ws.min_column + 20, values_only=True), 1):
            vals = [str(v).strip().lower() if v else '' for v in row]
            if all(k in vals for k in keywords):
                return vals, [vals.index(k) for k in keywords], i
        return None, None, None

    if 'Droplist' in wb.sheetnames:
        ws = wb['Droplist']
        mc = ws.min_column
        if mc < 1:
            mc = 1
        week_count = 0
        for row in ws.iter_rows(min_row=2, min_col=mc, max_col=mc + 15, values_only=True):
            vals = list(row)
            date_raw = str(vals[0]).strip() if len(vals) > 0 and vals[0] else ''
            week_raw = vals[1] if len(vals) > 1 else None
            if isinstance(week_raw, (int, float)) and week_raw > 0:
                if not Week.query.filter_by(week_number=int(week_raw)).first():
                    db.session.add(Week(week_number=int(week_raw), date_range=date_raw))
                    week_count += 1
        store_names = set()
        for row in ws.iter_rows(min_row=2, min_col=mc, max_col=mc + 15, values_only=True):
            vals = list(row)
            store_val = str(vals[3]).strip() if len(vals) > 3 and vals[3] else ''
            if store_val:
                store_names.add(store_val)
        for name in store_names:
            if not Store.query.filter_by(name=name).first():
                db.session.add(Store(name=name))
        rows_data = []
        for row in ws.iter_rows(min_row=5, min_col=mc, max_col=mc + 15, values_only=True):
            vals = list(row)
            u = str(vals[7]).strip() if len(vals) > 7 and vals[7] else ''
            w = str(vals[8]).strip() if len(vals) > 8 and vals[8] else ''
            s = str(vals[9]).strip() if len(vals) > 9 and vals[9] else ''
            if u or s:
                rows_data.append((u, w, s))
        seen_units = set()
        for u, w, s in rows_data:
            if u and u not in seen_units:
                seen_units.add(u)
                if not Unit.query.filter_by(name=u).first():
                    db.session.add(Unit(name=u))
            if s and not Section.query.filter_by(name=s).first():
                db.session.add(Section(name=s))
        db.session.commit()
        for u, w, s in rows_data:
            if u and w:
                unit = Unit.query.filter_by(name=u).first()
                if unit and not Ward.query.filter_by(name=w, unit_id=unit.id).first():
                    db.session.add(Ward(name=w, unit_id=unit.id))
        db.session.commit()
        results.append(f'{week_count} weeks, {len(store_names)} stores, {len(set(r[0] for r in rows_data if r[0]))} units, {len(set(r[2] for r in rows_data if r[2]))} sections seeded')

    if 'Master List2' in wb.sheetnames:
        ws = wb['Master List2']
        header, idxs, hrow = _find_header(ws, ['code'])
        if header:
            code_idx = idxs[0]
            name_idx = header.index('item') if 'item' in header else code_idx + 1
            store_idx = header.index('store') if 'store' in header else code_idx + 2
            stock_idx = header.index('stock quantity') if 'stock quantity' in header else None
            cl_idx = header.index('critical level') if 'critical level' in header else None
            need = [i for i in (code_idx, name_idx, store_idx) if i is not None]
            all_cols = [i for i in (code_idx, name_idx, store_idx, stock_idx, cl_idx) if i is not None]
            max_col = max(all_cols) + 1 + ws.min_column
            item_count = 0
            for row in ws.iter_rows(min_row=hrow + 1, min_col=ws.min_column, max_col=max_col, values_only=True):
                vals = [str(v).strip() if v is not None else '' for v in row]
                if len(vals) <= max(need):
                    continue
                code = vals[code_idx]
                name = vals[name_idx] if name_idx is not None and len(vals) > name_idx else ''
                store_name = vals[store_idx] if store_idx is not None and len(vals) > store_idx else ''
                if not code or not name or not store_name:
                    continue
                store = Store.query.filter_by(name=store_name).first()
                if not store:
                    store = Store(name=store_name)
                    db.session.add(store)
                    db.session.commit()
                if not Item.query.filter_by(code=code).first():
                    item = Item(code=code, name=name, store_id=store.id)
                    if stock_idx is not None and len(vals) > stock_idx and vals[stock_idx]:
                        try: item.stock_quantity = int(float(vals[stock_idx]))
                        except: pass
                    if cl_idx is not None and len(vals) > cl_idx and vals[cl_idx]:
                        try: item.critical_level = int(float(vals[cl_idx]))
                        except: pass
                    db.session.add(item)
                    item_count += 1
            db.session.commit()
            results.append(f'{item_count} items imported')
        else:
            results.append('Master List2: header not found')

    if 'Report' in wb.sheetnames:
        ws = wb['Report']
        header, idxs, hrow = _find_header(ws, ['code', 'status'])
        if header:
            code_idx = idxs[0]
            status_idx = idxs[1]
            name_idx = header.index('item') if 'item' in header else None
            store_idx = header.index('store') if 'store' in header else None
            week_idx = header.index('week') if 'week' in header else None
            unit_idx = header.index('unit') if 'unit' in header else None
            ward_idx = header.index('ward') if 'ward' in header else None
            section_idx = header.index('section') if 'section' in header else None
            need = [i for i in (code_idx, status_idx) if i is not None]
            all_cols = [i for i in (code_idx, status_idx, name_idx, store_idx, week_idx, unit_idx, ward_idx, section_idx) if i is not None]
            max_col = max(all_cols) + 1 + ws.min_column
            import re
            report_count = 0
            errors = []
            for i, row in enumerate(ws.iter_rows(min_row=hrow + 1, min_col=ws.min_column, max_col=max_col, values_only=True), start=hrow + 1):
                vals = [str(v).strip() if v is not None else '' for v in row]
                if len(vals) <= max(need):
                    continue
                code = vals[code_idx]
                status_raw = vals[status_idx]
                if not code or not status_raw:
                    continue
                section_name = vals[section_idx] if section_idx is not None and len(vals) > section_idx else ''
                unit_name = vals[unit_idx] if unit_idx is not None and len(vals) > unit_idx else ''
                ward_name = vals[ward_idx] if ward_idx is not None and len(vals) > ward_idx else ''
                item_name = vals[name_idx] if name_idx is not None and len(vals) > name_idx else ''
                store_name = vals[store_idx] if store_idx is not None and len(vals) > store_idx else ''
                week_raw = vals[week_idx] if week_idx is not None and len(vals) > week_idx else ''
                item = Item.query.filter_by(code=code).first()
                if not item:
                    errors.append(f'Row {i}: Item code "{code}" not found')
                    continue
                nums = re.findall(r'\d+', str(week_raw))
                week_num = int(nums[0]) if nums else None
                week = Week.query.filter_by(week_number=week_num).first() if week_num else None
                if not week and week_num:
                    week = Week(week_number=week_num, date_range=str(week_raw))
                    db.session.add(week)
                    db.session.commit()
                if not week:
                    errors.append(f'Row {i}: Week not found ({week_raw})')
                    continue
                unit = Unit.query.filter_by(name=unit_name).first() if unit_name else None
                ward = Ward.query.filter_by(name=ward_name).first() if ward_name else None
                section = Section.query.filter_by(name=section_name).first() if section_name else None
                if not unit and unit_name:
                    unit = Unit(name=unit_name)
                    db.session.add(unit)
                    db.session.commit()
                if not ward and unit and ward_name:
                    ward = Ward(name=ward_name, unit_id=unit.id)
                    db.session.add(ward)
                    db.session.commit()
                if not section and section_name:
                    section = Section(name=section_name)
                    db.session.add(section)
                    db.session.commit()
                s = 'Shortage' if status_raw.lower() == 'shortage' else 'Not available'
                if not Report.query.filter_by(item_id=item.id, week_id=week.id, status=s).first():
                    db.session.add(Report(
                        item_id=item.id, week_id=week.id,
                        unit_id=unit.id if unit else 1,
                        ward_id=ward.id if ward else 1,
                        section_id=section.id if section else 1,
                        status=s
                    ))
                    report_count += 1
            db.session.commit()
            msg = f'{report_count} reports imported'
            if errors:
                msg += f' ({len(errors)} errors: {errors[0]}{"..." if len(errors)>1 else ""})'
            results.append(msg)
        else:
            results.append('Report: header not found')

    flash('Import complete: ' + ' | '.join(results), 'success')
    return redirect(url_for('dashboard'))


@app.route('/settings/clear', methods=['POST'])
def clear_data():
    Report.query.delete()
    Item.query.delete()
    Store.query.delete()
    Week.query.delete()
    Unit.query.delete()
    Ward.query.delete()
    Section.query.delete()
    db.session.commit()
    _log_activity('System', 'Cleared all supply data')
    flash('All data cleared', 'success')
    return redirect(url_for('settings'))


@app.route('/download/template/equipment')
def download_equipment_template():
    import openpyxl
    from io import BytesIO
    from flask import send_file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Equipment List'
    ws.append(['Biomedical', 'Name', 'Status', 'Location', 'Remarks'])
    ws.append(['EQM-001', 'Patient Monitor', 'WORKING', 'ER Room 3', ''])
    ws.append(['DEFIB-001', 'Defibrillator', 'WORKING', 'ER Code Cart', ''])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='PULSE_equipment_template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/download/template/items')
def download_items_template():
    import openpyxl
    from io import BytesIO
    from flask import send_file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Master List2'
    ws.append(['Code', 'Item', 'Store', 'Stock Quantity', 'Critical Level'])
    ws.append(['S-001', 'Sample Item', 'Pharmacy', 50, 10])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='PULSE_items_template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/export/reports/xlsx')
def export_reports_xlsx():
    store_id = request.args.get('store_id', type=int)
    status_filter = request.args.get('status')
    week_id = request.args.get('week_id', type=int)
    query = Report.query
    if store_id: query = query.join(Item).filter(Item.store_id == store_id)
    if status_filter: query = query.filter(Report.status == status_filter)
    if week_id: query = query.filter(Report.week_id == week_id)
    reports = query.order_by(Report.created_at.desc()).all()
    import openpyxl
    from io import BytesIO
    from flask import send_file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Reports'
    headers = ['Item Code', 'Item Name', 'Store', 'Week', 'Unit', 'Ward', 'Section', 'Status', 'Reported']
    ws.append(headers)
    _style_xl(ws, headers, [12, 24, 14, 10, 14, 14, 14, 12, 14])
    for r in reports:
        ws.append([r.item.code, r.item.name, r.item.store.name, f'Week {r.week.week_number}',
                   r.unit.name, r.ward.name, r.section.name, r.status,
                   r.created_at.strftime('%Y-%m-%d') if r.created_at else ''])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='PULSE_reports.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/export/reports')
def export_reports():
    import openpyxl
    from io import BytesIO
    from flask import send_file

    store_id = request.args.get('store_id', type=int)
    status_filter = request.args.get('status')
    week_id = request.args.get('week_id', type=int)

    query = Report.query
    if store_id:
        query = query.join(Item).filter(Item.store_id == store_id)
    if status_filter:
        query = query.filter(Report.status == status_filter)
    if week_id:
        query = query.filter(Report.week_id == week_id)
    query = query.order_by(Report.created_at.desc())
    reports = query.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Reports'
    headers = ['Item Code', 'Item Name', 'Store', 'Week', 'Unit', 'Ward', 'Section', 'Status', 'Reported']
    ws.append(headers)
    _style_xl(ws, headers, [12, 24, 14, 10, 14, 14, 14, 12, 14])
    for r in reports:
        ws.append([r.item.code, r.item.name, r.item.store.name, f'Week {r.week.week_number}',
                   r.unit.name, r.ward.name, r.section.name, r.status,
                   r.created_at.strftime('%Y-%m-%d')])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='PULSE_reports.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/export/shortage')
def export_shortage():
    import openpyxl
    from io import BytesIO
    from flask import send_file

    store_id = request.args.get('store_id', type=int)
    from_date = request.args.get('from')
    to_date = request.args.get('to')
    query = db.session.query(Item.code, Item.name, Store.name.label('store'),
                             func.count(Report.id).label('total')
                             ).select_from(Report).join(Item).join(Store).filter(Report.status == 'Shortage')
    if store_id:
        query = query.filter(Item.store_id == store_id)
    if from_date:
        fd = datetime.strptime(from_date, '%Y-%m-%d')
        query = query.filter(Report.created_at >= fd)
    if to_date:
        td = datetime.strptime(to_date, '%Y-%m-%d') + timedelta(days=1)
        query = query.filter(Report.created_at < td)
    data = query.group_by(Item.id, Item.code, Item.name, Store.name).order_by(func.count(Report.id).desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Shortage'
    headers = ['Code', 'Item', 'Store', 'Times Reported']
    ws.append(headers)
    _style_xl(ws, headers, [12, 28, 16, 16])
    for code, name, store, total in data:
        ws.append([code, name, store, total])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='PULSE_shortage.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/export/una')
def export_una():
    import openpyxl
    from io import BytesIO
    from flask import send_file

    store_id = request.args.get('store_id', type=int)
    from_date = request.args.get('from')
    to_date = request.args.get('to')
    query = db.session.query(Item.code, Item.name, Store.name.label('store'),
                             func.count(Report.id).label('total')
                             ).select_from(Report).join(Item).join(Store).filter(Report.status == 'Not available')
    if store_id:
        query = query.filter(Item.store_id == store_id)
    if from_date:
        fd = datetime.strptime(from_date, '%Y-%m-%d')
        query = query.filter(Report.created_at >= fd)
    if to_date:
        td = datetime.strptime(to_date, '%Y-%m-%d') + timedelta(days=1)
        query = query.filter(Report.created_at < td)
    data = query.group_by(Item.id, Item.code, Item.name, Store.name).order_by(func.count(Report.id).desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Not Available'
    headers = ['Code', 'Item', 'Store', 'Times Reported']
    ws.append(headers)
    _style_xl(ws, headers, [12, 28, 16, 16])
    for code, name, store, total in data:
        ws.append([code, name, store, total])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='PULSE_una.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



@app.route('/search')
def search():
    q = request.args.get('q', '').strip()
    scope = request.args.get('scope', 'supply')
    items = []
    reports = []
    equipment = []
    if q:
        if scope in ('supply', 'all'):
            items = Item.query.filter(
                Item.code.ilike(f'%{q}%') | Item.name.ilike(f'%{q}%')
            ).order_by(Item.code).limit(30).all()
            reports = Report.query.join(Item).filter(
                Item.code.ilike(f'%{q}%') | Item.name.ilike(f'%{q}%')
            ).order_by(Report.created_at.desc()).limit(30).all()
        if scope in ('equipment', 'all'):
            equipment = Equipment.query.filter(
                Equipment.biomed.ilike(f'%{q}%') |
                Equipment.name.ilike(f'%{q}%') |
                Equipment.location.ilike(f'%{q}%')
            ).order_by(Equipment.biomed).limit(30).all()
    return render_template('search.html', q=q, scope=scope, items=items, reports=reports, equipment=equipment)


# ── Equipment Module ──

@app.route('/equipment/dashboard')
@app.route('/equipment')
def equipment_dashboard():
    total = Equipment.query.count()
    working = Equipment.query.filter_by(status='WORKING').count()
    missing = Equipment.query.filter_by(status='MISSING').count()
    loaned = Equipment.query.filter_by(status='LOANED').count()
    maintenance = Equipment.query.filter_by(status='MAINTENANCE').count()
    condemned = Equipment.query.filter_by(status='CONDEMNED').count()
    today = datetime.now().date()
    overdue_q = Equipment.query.filter(
        Equipment.next_service_date.isnot(None),
        Equipment.next_service_date < today
    )
    overdue = overdue_q.count()
    overdue_list = overdue_q.order_by(Equipment.next_service_date).limit(8).all()
    due_soon_q = Equipment.query.filter(
        Equipment.next_service_date.isnot(None),
        Equipment.next_service_date >= today,
        Equipment.next_service_date <= today + timedelta(days=30)
    )
    due_soon = due_soon_q.count()
    due_soon_list = due_soon_q.order_by(Equipment.next_service_date).limit(8).all()
    missing_list = Equipment.query.filter_by(status='MISSING').order_by(Equipment.name).limit(5).all()
    recent = EquipmentLog.query.order_by(EquipmentLog.created_at.desc()).limit(10).all()
    checked_out = EquipmentMovement.query.filter_by(status='OUT').count()
    recent_out = EquipmentMovement.query.filter_by(status='OUT').order_by(EquipmentMovement.checked_out_at.desc()).limit(6).all()
    checkout_today = EquipmentMovement.query.filter(
        func.date(EquipmentMovement.checked_out_at) == today
    ).count()
    overdue_checkouts = EquipmentMovement.query.filter(
        EquipmentMovement.status == 'OUT',
        EquipmentMovement.checked_out_at < today - timedelta(days=7)
    ).count()
    from sqlalchemy import desc
    top_checked = db.session.query(
        EquipmentMovement.equipment_id, func.count(EquipmentMovement.id).label('cnt')
    ).group_by(EquipmentMovement.equipment_id).order_by(desc('cnt')).limit(5).all()
    top_equipment = []
    for eq_id, cnt in top_checked:
        eq = db.session.get(Equipment, eq_id)
        if eq:
            top_equipment.append((eq, cnt))
    return render_template('equipment/dashboard.html',
        total=total, working=working, missing=missing,
        checked_out=checked_out,
        loaned=loaned, maintenance=maintenance, condemned=condemned,
        overdue=overdue, overdue_list=overdue_list,
        due_soon=due_soon, due_soon_list=due_soon_list,
        missing_list=missing_list,
        recent=recent, recent_out=recent_out,
        checkout_today=checkout_today,
        overdue_checkouts=overdue_checkouts,
        top_equipment=top_equipment)


@app.route('/equipment/list')
def equipment_list():
    q = request.args.get('q', '').strip()
    status_filter = request.args.get('status', '')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    eq = Equipment.query
    if q:
        eq = eq.filter(
            Equipment.biomed.ilike(f'%{q}%') |
            Equipment.name.ilike(f'%{q}%') |
            Equipment.location.ilike(f'%{q}%')
        )
    if status_filter:
        eq = eq.filter(Equipment.status == status_filter)
    eq = eq.order_by(Equipment.biomed)
    pagination = eq.paginate(page=page, per_page=per_page, error_out=False)
    # Get all currently checked-out equipment IDs
    active_out_ids = set(m.equipment_id for m in EquipmentMovement.query.filter_by(status='OUT').all())
    return render_template('equipment/list.html',
        equipment=pagination.items, pagination=pagination,
        active_out_ids=active_out_ids, per_page=per_page)


@app.route('/equipment/add', methods=['POST'])
def equipment_add():
    biomed = request.form.get('biomed', '').strip().upper()
    name = request.form.get('name', '').strip()
    if not biomed or not name:
        flash('Biomedical number and name are required', 'danger')
        return redirect(url_for('equipment_list'))
    if Equipment.query.filter_by(biomed=biomed).first():
        flash(f'Equipment "{biomed}" already exists', 'danger')
        return redirect(url_for('equipment_list'))
    eq = Equipment(
        biomed=biomed, name=name,
        status=request.form.get('status', 'WORKING'),
        location=request.form.get('location', '').strip(),
        remarks=request.form.get('remarks', '').strip(),
    )
    ls = request.form.get('last_service_date', '')
    ns = request.form.get('next_service_date', '')
    if ls:
        try: eq.last_service_date = datetime.strptime(ls, '%Y-%m-%d')
        except: pass
    if ns:
        try: eq.next_service_date = datetime.strptime(ns, '%Y-%m-%d')
        except: pass
    db.session.add(eq)
    db.session.commit()
    _log_equipment(eq.id, 'ADDED', 'System', 'Equipment added to database')
    flash(f'Equipment "{biomed}" added', 'success')
    return redirect(url_for('equipment_list'))


@app.route('/equipment/<int:id>/edit', methods=['POST'])
def equipment_edit(id):
    eq = db.session.get(Equipment, id)
    if not eq:
        flash('Equipment not found', 'danger')
        return redirect(url_for('equipment_list'))
    eq.name = request.form.get('name', eq.name).strip()
    eq.status = request.form.get('status', eq.status)
    eq.location = request.form.get('location', eq.location).strip()
    eq.remarks = request.form.get('remarks', eq.remarks).strip()
    ls = request.form.get('last_service_date', '')
    ns = request.form.get('next_service_date', '')
    if ls:
        try: eq.last_service_date = datetime.strptime(ls, '%Y-%m-%d')
        except: pass
    else:
        eq.last_service_date = None
    if ns:
        try: eq.next_service_date = datetime.strptime(ns, '%Y-%m-%d')
        except: pass
    else:
        eq.next_service_date = None
    db.session.commit()
    _log_equipment(eq.id, 'UPDATED', 'System', 'Equipment details updated')
    flash('Equipment updated', 'success')
    return redirect(url_for('equipment_list'))


@app.route('/equipment/<int:id>/delete', methods=['POST'])
def equipment_delete(id):
    eq = db.session.get(Equipment, id)
    if eq:
        EquipmentLog.query.filter_by(equipment_id=id).delete()
        db.session.delete(eq)
        db.session.commit()
        flash('Equipment deleted', 'success')
    return redirect(url_for('equipment_list'))


@app.route('/equipment/<int:id>/scan', methods=['POST'])
def equipment_scan(id):
    eq = db.session.get(Equipment, id)
    if not eq:
        return jsonify({'error': 'Not found'}), 404
    scanned_by = request.form.get('scanned_by', 'Staff').strip()
    eq.status = 'WORKING'
    eq.last_scanned_date = datetime.now(timezone.utc)
    eq.last_scanned_by = scanned_by
    db.session.commit()
    _log_equipment(eq.id, 'SCANNED', scanned_by, 'Equipment scanned as WORKING')
    return jsonify({'ok': True, 'status': 'WORKING'})


@app.route('/equipment/<int:id>/status', methods=['POST'])
def equipment_set_status(id):
    eq = db.session.get(Equipment, id)
    if not eq:
        return jsonify({'error': 'Not found'}), 404
    new_status = request.form.get('status', '')
    valid = ['WORKING', 'MISSING', 'LOANED', 'MAINTENANCE', 'CONDEMNED']
    if new_status not in valid:
        return jsonify({'error': 'Invalid status'}), 400
    eq.status = new_status
    db.session.commit()
    _log_equipment(eq.id, 'STATUS_CHANGED', 'System', f'Status changed to {new_status}')
    return jsonify({'ok': True, 'status': new_status})


@app.route('/api/equipment')
def api_equipment():
    eqs = Equipment.query.order_by(Equipment.biomed).all()
    today = datetime.now().date()
    data = []
    for e in eqs:
        overdue = False
        due_soon = False
        if e.next_service_date:
            days = (e.next_service_date.date() - today).days if e.next_service_date else 999
            overdue = days < 0
            due_soon = 0 <= days <= 30
        data.append({
            'id': e.id, 'biomed': e.biomed, 'name': e.name,
            'status': e.status, 'location': e.location or '',
            'last_service': e.last_service_date.strftime('%Y-%m-%d') if e.last_service_date else '',
            'next_service': e.next_service_date.strftime('%Y-%m-%d') if e.next_service_date else '',
            'overdue': overdue, 'due_soon': due_soon,
        })
    return jsonify(data)


@app.route('/api/equipment/dashboard')
def api_equipment_dashboard():
    total = Equipment.query.count()
    working = Equipment.query.filter_by(status='WORKING').count()
    missing = Equipment.query.filter_by(status='MISSING').count()
    loaned = Equipment.query.filter_by(status='LOANED').count()
    maintenance = Equipment.query.filter_by(status='MAINTENANCE').count()
    condemned = Equipment.query.filter_by(status='CONDEMNED').count()
    today = datetime.now().date()
    overdue_eqs = Equipment.query.filter(
        Equipment.next_service_date.isnot(None),
        Equipment.next_service_date < today
    ).all()
    due_soon_eqs = Equipment.query.filter(
        Equipment.next_service_date.isnot(None),
        Equipment.next_service_date >= today,
        Equipment.next_service_date <= datetime(today.year, today.month, today.day) + timedelta(days=30)
    ).all()
    recent = EquipmentLog.query.order_by(EquipmentLog.created_at.desc()).limit(10).all()
    checked_out_count = EquipmentMovement.query.filter_by(status='OUT').count()
    return jsonify({
        'total': total, 'working': working, 'missing': missing,
        'checked_out': checked_out_count,
        'loaned': loaned, 'maintenance': maintenance, 'condemned': condemned,
        'overdue_count': len(overdue_eqs),
        'due_soon_count': len(due_soon_eqs),
        'overdue': [{'id': e.id, 'biomed': e.biomed, 'name': e.name,
                     'next_service': e.next_service_date.strftime('%Y-%m-%d') if e.next_service_date else ''} for e in overdue_eqs],
        'due_soon': [{'id': e.id, 'biomed': e.biomed, 'name': e.name,
                      'next_service': e.next_service_date.strftime('%Y-%m-%d') if e.next_service_date else ''} for e in due_soon_eqs],
        'recent': [{'id': r.id, 'equipment': r.equipment.biomed + ' - ' + r.equipment.name,
                    'action': r.action, 'scanned_by': r.scanned_by or '',
                    'date': r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else ''} for r in recent],
    })


@app.route('/equipment/service')
def equipment_service():
    today = datetime.now().date()
    overdue = Equipment.query.filter(
        Equipment.next_service_date.isnot(None),
        Equipment.next_service_date < today
    ).order_by(Equipment.next_service_date).all()
    due_soon = Equipment.query.filter(
        Equipment.next_service_date.isnot(None),
        Equipment.next_service_date >= today,
        Equipment.next_service_date <= datetime(today.year, today.month, today.day) + timedelta(days=30)
    ).order_by(Equipment.next_service_date).all()
    upcoming = Equipment.query.filter(
        Equipment.next_service_date.isnot(None),
        Equipment.next_service_date > today + timedelta(days=30)
    ).order_by(Equipment.next_service_date).limit(20).all()
    return render_template('equipment/service.html',
        overdue=overdue, due_soon=due_soon, upcoming=upcoming)


@app.route('/equipment/qr')
def equipment_qr():
    q = request.args.get('q', '').strip()
    eqs = Equipment.query.order_by(Equipment.biomed)
    if q:
        eqs = eqs.filter(
            Equipment.biomed.ilike(f'%{q}%') |
            Equipment.name.ilike(f'%{q}%')
        )
    eqs = eqs.all()
    return render_template('equipment/qr.html', equipment=eqs)


@app.route('/equipment/qr/image/<int:id>')
def equipment_qr_image(id):
    import qrcode
    from io import BytesIO
    eq = db.session.get(Equipment, id)
    if not eq:
        return render_template('404.html'), 404
    content = eq.biomed
    img = qrcode.make(content, box_size=12, border=2)
    buf = BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return Response(buf.getvalue(), mimetype='image/png')


@app.route('/equipment/qr/print', methods=['GET'])
def equipment_qr_print():
    ids = request.args.getlist('ids')
    eqs = Equipment.query.filter(Equipment.id.in_(ids)).order_by(Equipment.biomed).all() if ids else []
    return render_template('equipment/qr_print.html', equipment=eqs)


@app.route('/equipment/export')
def equipment_export():
    import openpyxl
    from io import BytesIO
    from flask import send_file
    eqs = Equipment.query.order_by(Equipment.biomed).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Equipment'
    headers = ['Biomedical #', 'Name', 'Status', 'Location', 'Last Service', 'Next Service', 'Remarks']
    ws.append(headers)
    _style_xl(ws, headers, [16, 28, 12, 18, 14, 14, 22])
    for e in eqs:
        ws.append([e.biomed, e.name, e.status, e.location or '',
            e.last_service_date.strftime('%Y-%m-%d') if e.last_service_date else '',
            e.next_service_date.strftime('%Y-%m-%d') if e.next_service_date else '',
            e.remarks or ''])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='PULSE_equipment.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/equipment/scan')
def equipment_fastscan():
    mode = request.args.get('mode', 'checkout')
    return render_template('equipment/scan.html', mode=mode)


@app.route('/equipment/<int:id>')
def equipment_detail(id):
    eq = db.session.get(Equipment, id)
    if not eq:
        flash('Equipment not found', 'danger')
        return redirect(url_for('equipment_list'))
    movements = EquipmentMovement.query.filter_by(equipment_id=id).order_by(EquipmentMovement.checked_out_at.desc()).all()
    logs = EquipmentLog.query.filter_by(equipment_id=id).order_by(EquipmentLog.created_at.desc()).all()
    current_out = EquipmentMovement.query.filter_by(equipment_id=id, status='OUT').first()
    active_out_ids = set(m.equipment_id for m in EquipmentMovement.query.filter_by(status='OUT').all())
    return render_template('equipment/detail.html', eq=eq,
        movements=movements, logs=logs, current_out=current_out,
        active_out_ids=active_out_ids)


@app.route('/equipment/<int:id>/checkout', methods=['POST'])
def equipment_checkout(id):
    eq = db.session.get(Equipment, id)
    if not eq:
        flash('Equipment not found', 'danger')
        return redirect(url_for('equipment_list'))
    active = EquipmentMovement.query.filter_by(equipment_id=id, status='OUT').first()
    if active:
        flash(f'Equipment is already checked out by {active.checked_out_by}', 'danger')
        return redirect(url_for('equipment_detail', id=id))
    checked_out_by = request.form.get('checked_out_by', '').strip().upper()
    destination = request.form.get('destination', '').strip()
    purpose = request.form.get('purpose', '').strip()
    source_location = request.form.get('source_location', eq.location or '').strip()
    if not checked_out_by or not destination:
        flash('Employee ID and destination are required', 'danger')
        return redirect(url_for('equipment_detail', id=id))
    staff = Staff.query.filter_by(employee_id=checked_out_by).first()
    if not staff:
        flash(f'Employee ID "{checked_out_by}" not found in staff list', 'danger')
        return redirect(url_for('equipment_detail', id=id))
    m = EquipmentMovement(
        equipment_id=id, checked_out_by=checked_out_by,
        source_location=source_location, destination=destination,
        purpose=purpose, status='OUT'
    )
    eq.status = 'LOANED'
    db.session.add(m)
    db.session.commit()
    _log_equipment(id, 'CHECKED_OUT', checked_out_by,
        f'To: {destination} | Purpose: {purpose} | By: {staff.full_name}')
    flash(f'{eq.biomed} checked out to {staff.full_name}', 'success')
    return redirect(url_for('equipment_detail', id=id))


@app.route('/equipment/<int:id>/checkin', methods=['POST'])
def equipment_checkin(id):
    m = EquipmentMovement.query.filter_by(equipment_id=id, status='OUT').first()
    if not m:
        flash('No active check-out found', 'danger')
        return redirect(url_for('equipment_detail', id=id))
    checked_in_by = request.form.get('checked_in_by', '').strip().upper()
    if checked_in_by:
        staff = Staff.query.filter_by(employee_id=checked_in_by).first()
        if not staff:
            flash(f'Employee ID "{checked_in_by}" not found in staff list', 'danger')
            return redirect(url_for('equipment_detail', id=id))
        m.checked_in_by = checked_in_by
    return_status = request.form.get('return_status', 'WORKING').strip().upper()
    if return_status not in ('WORKING', 'MAINTENANCE', 'CONDEMNED'):
        return_status = 'WORKING'
    return_notes = request.form.get('return_notes', '').strip()
    m.checked_in_at = datetime.now(timezone.utc)
    m.status = 'RETURNED'
    eq = db.session.get(Equipment, id)
    if eq:
        eq.status = return_status
    db.session.commit()
    log_notes = f'Returned from {m.destination}'
    if return_status == 'MAINTENANCE':
        log_notes += ' [Needs Repair]'
    elif return_status == 'CONDEMNED':
        log_notes += ' [CONDEMNED]'
    if return_notes:
        log_notes += f' — {return_notes}'
    _log_equipment(id, 'CHECKED_IN', checked_in_by or m.checked_out_by, log_notes)
    status_label = {'WORKING': 'checked in', 'MAINTENANCE': 'marked for repair', 'CONDEMNED': 'condemned'}
    flash(f'{eq.biomed} {status_label.get(return_status, "checked in")}', 'success')
    return redirect(url_for('equipment_detail', id=id))


@app.route('/equipment/tracking/checkin-all', methods=['POST'])
def equipment_checkin_all():
    ids = request.form.getlist('ids')
    checked_in_by = request.form.get('checked_in_by', '').strip().upper()
    if checked_in_by:
        staff = Staff.query.filter_by(employee_id=checked_in_by).first()
        if not staff:
            flash(f'Employee ID "{checked_in_by}" not found in staff list', 'danger')
            return redirect(url_for('equipment_tracking'))
    return_status = request.form.get('return_status', 'WORKING').strip().upper()
    if return_status not in ('WORKING', 'MAINTENANCE', 'CONDEMNED'):
        return_status = 'WORKING'
    return_notes = request.form.get('return_notes', '').strip()
    count = 0
    for eid in ids:
        m = EquipmentMovement.query.filter_by(equipment_id=int(eid), status='OUT').first()
        if m:
            if checked_in_by:
                m.checked_in_by = checked_in_by
            m.checked_in_at = datetime.now(timezone.utc)
            m.status = 'RETURNED'
            eq = db.session.get(Equipment, int(eid))
            if eq:
                eq.status = return_status
            log_notes = 'Bulk check-in'
            if return_status == 'MAINTENANCE':
                log_notes += ' [Needs Repair]'
            elif return_status == 'CONDEMNED':
                log_notes += ' [CONDEMNED]'
            if return_notes:
                log_notes += f' — {return_notes}'
            _log_equipment(int(eid), 'CHECKED_IN', checked_in_by or m.checked_out_by, log_notes)
            count += 1
    db.session.commit()
    status_label = {'WORKING': 'checked in', 'MAINTENANCE': 'marked for repair', 'CONDEMNED': 'condemned'}
    flash(f'{count} equipment {status_label.get(return_status, "checked in")}', 'success')
    return redirect(url_for('equipment_tracking'))


# ── Staff module ──

@app.route('/staff/list')
@admin_required
def staff_list():
    pending = Staff.query.filter_by(approved=False).order_by(Staff.created_at.desc()).all()
    approved = Staff.query.filter_by(approved=True).order_by(Staff.employee_id).all()
    return render_template('staff.html', pending=pending, approved=approved)


@app.route('/staff/add', methods=['POST'])
@admin_required
def staff_add():
    eid = request.form.get('employee_id', '').strip().upper()
    name = request.form.get('full_name', '').strip()
    dept = request.form.get('department', '').strip()
    if not eid or not name:
        flash('Employee ID and Full Name are required', 'danger')
        return redirect(url_for('staff_list'))
    if Staff.query.filter_by(employee_id=eid).first():
        flash(f'Employee ID "{eid}" already exists', 'danger')
        return redirect(url_for('staff_list'))
    password = request.form.get('password', '')
    role = request.form.get('role', 'staff')
    s = Staff(employee_id=eid, full_name=name, department=dept, role=role)
    if password:
        s.set_password(password)
    db.session.add(s)
    db.session.commit()
    _log_activity('Settings', f'Added staff: {eid} — {name}')
    flash(f'Staff {eid} ({name}) added', 'success')
    return redirect(url_for('staff_list'))


@app.route('/staff/set-password', methods=['POST'])
@admin_required
def staff_set_password():
    sid = request.form.get('staff_id', type=int)
    password = request.form.get('password', '')
    if not sid or not password:
        flash('Invalid request', 'danger')
        return redirect(url_for('staff_list'))
    s = db.session.get(Staff, sid)
    if not s:
        flash('Staff not found', 'danger')
        return redirect(url_for('staff_list'))
    s.set_password(password)
    db.session.commit()
    _log_activity('Settings', f'Set password for: {s.employee_id}')
    flash(f'Password set for {s.employee_id}', 'success')
    return redirect(url_for('staff_list'))


@app.route('/account/change-password', methods=['POST'])
def account_change_password():
    current_user = db.session.get(Staff, session.get('user_id')) if session.get('user_id') else None
    if not current_user:
        flash('Not logged in', 'danger')
        return redirect(url_for('login'))
    old_pw = request.form.get('old_password', '')
    new_pw = request.form.get('new_password', '')
    confirm_pw = request.form.get('confirm_password', '')
    if not current_user.check_password(old_pw):
        flash('Current password is incorrect', 'danger')
        return redirect(url_for('dashboard'))
    if len(new_pw) < 8:
        flash('New password must be at least 8 characters', 'danger')
        return redirect(url_for('dashboard'))
    if new_pw != confirm_pw:
        flash('New passwords do not match', 'danger')
        return redirect(url_for('dashboard'))
    current_user.set_password(new_pw)
    db.session.commit()
    _log_activity(current_user.employee_id, 'Password changed')
    flash('Password updated successfully', 'success')
    return redirect(url_for('dashboard'))


@app.route('/staff/<int:id>/approve', methods=['POST'])
@admin_required
def staff_approve(id):
    s = db.session.get(Staff, id)
    if not s:
        flash('Staff not found', 'danger')
        return redirect(url_for('staff_list'))
    s.approved = True
    role = request.form.get('role', '')
    if role in ('admin', 'staff'):
        s.role = role
    db.session.commit()
    _log_activity('Settings', f'Approved staff: {s.employee_id} — {s.full_name} as {s.role}')
    flash(f'{s.employee_id} approved as {s.role}', 'success')
    return redirect(url_for('staff_list'))


@app.route('/staff/<int:id>/reject', methods=['POST'])
@admin_required
def staff_reject(id):
    s = db.session.get(Staff, id)
    if not s:
        flash('Staff not found', 'danger')
        return redirect(url_for('staff_list'))
    eid = s.employee_id
    db.session.delete(s)
    db.session.commit()
    _log_activity('Settings', f'Rejected/deleted staff registration: {eid}')
    flash(f'{eid} registration rejected and removed', 'success')
    return redirect(url_for('staff_list'))


@app.route('/staff/<int:id>/delete', methods=['POST'])
@admin_required
def staff_delete(id):
    s = db.session.get(Staff, id)
    if not s:
        flash('Staff not found', 'danger')
        return redirect(url_for('staff_list'))
    eid = s.employee_id
    db.session.delete(s)
    db.session.commit()
    _log_activity('Settings', f'Deleted staff: {eid}')
    flash(f'Staff {eid} deleted', 'success')
    return redirect(url_for('staff_list'))


@app.route('/api/staff/search')
def staff_search():
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify([])
    results = Staff.query.filter(
        Staff.employee_id.ilike(f'%{q}%') |
        Staff.full_name.ilike(f'%{q}%')
    ).order_by(Staff.employee_id).limit(15).all()
    return jsonify([{
        'id': s.id, 'employee_id': s.employee_id,
        'full_name': s.full_name, 'department': s.department or ''
    } for s in results])


@app.route('/api/staff/validate', methods=['POST'])
def staff_validate():
    eid = request.form.get('employee_id', '').strip().upper()
    s = Staff.query.filter_by(employee_id=eid).first()
    if s:
        return jsonify({'valid': True, 'full_name': s.full_name, 'department': s.department or ''})
    return jsonify({'valid': False})


@app.route('/api/activity')
def api_activity():
    logs = ActivityLog.query.order_by(ActivityLog.created_at.desc()).limit(50).all()
    return jsonify([{
        'id': l.id, 'user': l.user, 'action': l.action,
        'details': l.details,
        'date': l.created_at.strftime('%Y-%m-%d %H:%M') if l.created_at else ''
    } for l in logs])


@app.route('/equipment/export/log')
def equipment_export_log():
    import openpyxl
    from io import BytesIO
    from flask import send_file
    from datetime import datetime as dt
    eq_id = request.args.get('equipment_id', type=int)
    date_from = request.args.get('from', '')
    date_to = request.args.get('to', '')
    q = EquipmentLog.query
    if eq_id:
        q = q.filter(EquipmentLog.equipment_id == eq_id)
    if date_from:
        try: q = q.filter(EquipmentLog.created_at >= dt.strptime(date_from, '%Y-%m-%d'))
        except: pass
    if date_to:
        try: q = q.filter(EquipmentLog.created_at <= dt.strptime(date_to, '%Y-%m-%d') + timedelta(days=1))
        except: pass
    logs = q.order_by(EquipmentLog.created_at.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'AuditLog'
    headers = ['ID', 'Equipment', 'Biomed', 'Action', 'Scanned By', 'Notes', 'Date']
    ws.append(headers)
    _style_xl(ws, headers, [8, 28, 16, 12, 16, 24, 18])
    for l in logs:
        ws.append([l.id, l.equipment.name, l.equipment.biomed,
            l.action, l.scanned_by or '', l.notes or '',
            l.created_at.strftime('%Y-%m-%d %H:%M') if l.created_at else ''])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='PULSE_audit_log.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/equipment/tracking')
def equipment_tracking():
    q = request.args.get('q', '').strip()
    location_filter = request.args.get('location', '').strip()
    query = EquipmentMovement.query.filter_by(status='OUT').join(Equipment)
    if q:
        query = query.filter(
            Equipment.biomed.ilike(f'%{q}%') |
            Equipment.name.ilike(f'%{q}%') |
            EquipmentMovement.checked_out_by.ilike(f'%{q}%') |
            EquipmentMovement.destination.ilike(f'%{q}%')
        )
    if location_filter:
        query = query.filter(EquipmentMovement.destination.ilike(f'%{location_filter}%'))
    movements = query.order_by(EquipmentMovement.checked_out_at.desc()).all()
    checked_out_count = len(movements)
    # unique destination rooms
    destinations = db.session.query(EquipmentMovement.destination).filter(
        EquipmentMovement.status == 'OUT'
    ).distinct().order_by(EquipmentMovement.destination).all()
    return render_template('equipment/tracking.html',
        movements=movements, checked_out_count=checked_out_count,
        destinations=[d[0] for d in destinations if d[0]])


@app.route('/equipment/import', methods=['POST'])
def equipment_import():
    import openpyxl
    file = request.files.get('file')
    if not file:
        flash('No file selected', 'danger')
        return redirect(url_for('equipment_list'))
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        header = None
        hrow = 0
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, max_col=20, values_only=True), 1):
            vals = [str(v).strip().lower() if v else '' for v in row]
            if 'biomed' in vals or 'biomedical' in vals:
                header = vals
                hrow = i
                break
        if not header:
            flash('Could not find header row with "Biomedical" column', 'danger')
            return redirect(url_for('equipment_list'))
        biomed_idx = header.index('biomed') if 'biomed' in header else header.index('biomedical')
        name_idx = header.index('name') if 'name' in header else (biomed_idx + 1)
        status_idx = header.index('status') if 'status' in header else None
        location_idx = header.index('location') if 'location' in header else None
        remarks_idx = header.index('remarks') if 'remarks' in header else None
        need = [i for i in (biomed_idx, name_idx) if i is not None]
        all_cols = [i for i in (biomed_idx, name_idx, status_idx, location_idx, remarks_idx) if i is not None]
        max_col = max(all_cols) + 2
        count = 0
        for row in ws.iter_rows(min_row=hrow + 1, min_col=1, max_col=max_col, values_only=True):
            vals = [str(v).strip() if v is not None else '' for v in row]
            if len(vals) <= max(need):
                continue
            biomed = vals[biomed_idx].upper()
            name = vals[name_idx]
            if not biomed or not name:
                continue
            if Equipment.query.filter_by(biomed=biomed).first():
                continue
            eq = Equipment(biomed=biomed, name=name)
            if status_idx is not None and len(vals) > status_idx and vals[status_idx]:
                s = vals[status_idx].upper()
                if s in ('WORKING','MISSING','LOANED','MAINTENANCE','CONDEMNED'):
                    eq.status = s
            if location_idx is not None and len(vals) > location_idx:
                eq.location = vals[location_idx]
            if remarks_idx is not None and len(vals) > remarks_idx:
                eq.remarks = vals[remarks_idx]
            db.session.add(eq)
            count += 1
        db.session.commit()
        flash(f'{count} equipment imported successfully', 'success')
    except Exception as e:
        flash(f'Import failed: {e}', 'danger')
    return redirect(url_for('equipment_list'))


@app.route('/api/offline/replay', methods=['POST'])
def api_offline_replay():
    try:
        data = request.get_json(force=True) or {}
        action = data.get('action', '')
        if action == 'report':
            item_id = data.get('item_id')
            if item_id:
                current_user = db.session.get(Staff, session.get('user_id')) if session.get('user_id') else None
                unit_id = current_user.default_unit_id if current_user and current_user.default_unit_id else 1
                ward_id = current_user.default_ward_id if current_user and current_user.default_ward_id else 1
                week = Week.query.filter_by(week_number=datetime.now(timezone.utc).isocalendar()[1]).first()
                if week:
                    section = Section.query.first()
                    r = Report(item_id=int(item_id), week_id=week.id, unit_id=int(unit_id), ward_id=int(ward_id), section_id=section.id if section else 1, status='Shortage')
                    db.session.add(r); db.session.commit()
        return jsonify({'ok': True})
    except Exception:
        return jsonify({'ok': False}), 500


@app.route('/api/equipment/fastscan/lookup', methods=['POST'])
@csrf.exempt
def equipment_fastscan_lookup():
    code = request.form.get('code', '').strip().upper()
    if not code:
        return jsonify({'found': False})
    eq = Equipment.query.filter_by(biomed=code).first()
    if eq:
        return jsonify({'found': True, 'id': eq.id, 'biomed': eq.biomed,
                        'name': eq.name, 'status': eq.status, 'location': eq.location or ''})
    return jsonify({'found': False})


@app.route('/api/equipment/fastscan/finish', methods=['POST'])
@csrf.exempt
def equipment_fastscan_finish():
    ids = request.form.getlist('ids')
    mode = request.form.get('mode', 'checkout')
    scanned_by = request.form.get('scanned_by', '').strip().upper()
    if scanned_by:
        staff = Staff.query.filter_by(employee_id=scanned_by).first()
        if not staff:
            flash(f'Employee ID "{scanned_by}" not found in staff list', 'danger')
            return redirect(url_for('equipment_fastscan'))

    if mode == 'inspection':
        statuses = request.form.getlist('statuses')
        count = 0
        for i, eid in enumerate(ids):
            eq = db.session.get(Equipment, int(eid))
            if not eq:
                continue
            new_status = statuses[i] if i < len(statuses) else 'WORKING'
            old_status = eq.status
            eq.status = new_status
            eq.last_scanned_date = datetime.now(timezone.utc)
            eq.last_scanned_by = scanned_by
            _log_equipment(eq.id, 'STATUS_CHANGED', scanned_by,
                f'Inspection scan: {old_status} → {new_status}')
            count += 1
        db.session.commit()
        flash(f'{count} equipment updated via inspection', 'success')
        return redirect(url_for('equipment_fastscan', mode='inspection'))

    # Checkout mode
    destination = request.form.get('destination', '').strip()
    purpose = request.form.get('purpose', '').strip()
    checked_out = 0
    skipped = 0
    for eid in ids:
        eq = db.session.get(Equipment, int(eid))
        if not eq or eq.status == 'CONDEMNED':
            skipped += 1
            continue
        active = EquipmentMovement.query.filter_by(equipment_id=eq.id, status='OUT').first()
        if active:
            skipped += 1
            continue
        m = EquipmentMovement(
            equipment_id=eq.id, checked_out_by=scanned_by,
            source_location=eq.location or '', destination=destination,
            purpose=purpose, status='OUT'
        )
        eq.status = 'LOANED'
        eq.last_scanned_date = datetime.now(timezone.utc)
        eq.last_scanned_by = scanned_by
        db.session.add(m)
        _log_equipment(eq.id, 'CHECKED_OUT', scanned_by,
            f'Fast scan → To: {destination} | Purpose: {purpose}')
        checked_out += 1
    db.session.commit()
    msg = f'{checked_out} equipment checked out to {destination}'
    if skipped:
        msg += f' ({skipped} skipped — condemned or already out)'
    flash(msg, 'success')
    return redirect(url_for('equipment_fastscan'))


def _log_equipment(equipment_id, action, scanned_by, notes):
    log = EquipmentLog(equipment_id=equipment_id, action=action,
                       scanned_by=scanned_by, notes=notes)
    db.session.add(log)
    db.session.commit()


@app.errorhandler(403)
def forbidden(e):
    return render_template('403.html'), 403


@app.errorhandler(429)
def rate_limited(e):
    flash('Too many attempts. Please wait a minute.', 'danger')
    return redirect(url_for('login'))

@app.errorhandler(404)
def not_found(e):
    return render_template('404.html'), 404


@app.errorhandler(500)
def server_error(e):
    app.logger.error(f'500 error: {e}')
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Internal server error'}), 500
    return render_template('500.html'), 500


# ─── Activity Log Viewer ─────────────────────────────────────────────
@app.route('/activity-log')
@login_required
def activity_log():
    page = request.args.get('page', 1, type=int)
    q = request.args.get('q', '').strip()
    action = request.args.get('action', '').strip()
    query = ActivityLog.query
    if q:
        query = query.filter(
            ActivityLog.user.ilike(f'%{q}%') |
            ActivityLog.action.ilike(f'%{q}%') |
            ActivityLog.details.ilike(f'%{q}%')
        )
    if action:
        query = query.filter(ActivityLog.action == action)
    from_date = request.args.get('from')
    to_date = request.args.get('to')
    if from_date:
        fd = datetime.strptime(from_date, '%Y-%m-%d')
        query = query.filter(ActivityLog.created_at >= fd)
    if to_date:
        td = datetime.strptime(to_date, '%Y-%m-%d') + timedelta(days=1)
        query = query.filter(ActivityLog.created_at < td)
    query = query.order_by(ActivityLog.created_at.desc())
    pagination = query.paginate(page=page, per_page=50, error_out=False)
    actions = db.session.query(ActivityLog.action).distinct().order_by(ActivityLog.action).all()
    return render_template('activity_log.html',
        logs=pagination.items, pagination=pagination,
        actions=[a[0] for a in actions])


# ─── Database Backup ─────────────────────────────────────────────────
@app.route('/backup')
@admin_required
def backup_db():
    from io import BytesIO
    db_path = app.config['SQLALCHEMY_DATABASE_URI'].replace('sqlite:///', '')
    # make a copy to avoid locking issues
    backup_path = db_path + '.backup'
    shutil.copy2(db_path, backup_path)
    with open(backup_path, 'rb') as f:
        data = f.read()
    os.remove(backup_path)
    buf = BytesIO(data)
    from flask import send_file
    return send_file(buf, as_attachment=True,
        download_name=f'PULSE_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db',
        mimetype='application/octet-stream')


@app.route('/uploads/themes/<filename>')
def serve_theme_image(filename):
    return send_from_directory(THEME_UPLOAD_DIR, filename)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False)
